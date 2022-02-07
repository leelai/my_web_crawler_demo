import openpyxl
import easygui

pos_name = 'A'
pos_gender = 'B'
pos_birth = 'C'
pos_id = 'E'

id_str = "身份證"
name_str = "姓名"
birth_str= "生日"
birth2_str= "生日2"
year_ = "年"
month= "月"
day= "日"

def getUsers():
    myFile = easygui.fileopenbox()
    wb_obj = openpyxl.load_workbook(myFile)
    sheet = wb_obj.active

    store_list = []
    # total = int(sheet.max_row/3) - 1
    totalLine = int(sheet.max_row)
    # x = range(total)
    y = range(totalLine)
    for n in y:
        current_line = n + 1
        pos = pos_gender + str(current_line)
        name = sheet[pos].value
        if name is None:
            continue
        name = name.strip()
        if (name.find('男') == -1 & name.find('女') == -1):
            continue
        pos = pos_name + str(current_line)
        store_details = {id_str: None, name_str: None, birth_str: None, birth2_str:None, year_: None, month: None, day: None}
        name = sheet[pos].value
        if name is None:
            continue
        start = name.index('\n')
        name = name[start+1:20].strip()
        pos = pos_birth + str(current_line)
        birth = sheet[pos].value.replace('/','')
        if birth.index('0') == 0 :
            birth = birth[1:20].strip()

        birth2 = sheet[pos].value
        if birth2.index('0') == 0 :
            birth2 = birth2[1:20].strip()
        birth2 = birth2.split("/")

        store_details[year_] = birth2[0]
        store_details[month] = birth2[1]
        store_details[day] = birth2[2]
        year = int(birth2[0]) + 1911
        birth2 = str(year) + birth2[1] + birth2[2]
        pos = pos_id + str(current_line)
        id = sheet[pos].value.strip()
        if len(id) != 10:
            continue

        store_details[id_str] = id
        store_details[name_str] = name
        store_details[birth_str] = birth
        store_details[birth2_str] = birth2

        store_list.append(store_details)
    return store_list